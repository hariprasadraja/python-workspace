# Spreadsheet manipulation

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook(
    '/home/hariprasad/python/python-workspace/MoshTutorial/resources/transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = 0.0
    if cell.value is not None:
        corrected_price = cell.value * 0.9

    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = "$" + str(corrected_price)


values = Reference(
    sheet,
    min_row=2,
    max_row=sheet.max_row,
    min_col=4,
    max_col=4,
)

# XXX: chart is not working properly in wps office. test it with ms-office.
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('/home/hariprasad/python/python-workspace/MoshTutorial/resources/transactions.xlsx')
print("Spread sheet updated..")
